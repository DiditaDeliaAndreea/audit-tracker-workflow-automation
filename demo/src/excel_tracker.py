from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


BASE_DIR = Path(__file__).resolve().parents[1]
TRACKER_PATH = BASE_DIR / "output" / "audit_tracker_demo.xlsx"


def open_tracker():
    """Open the audit tracker workbook."""

    if not TRACKER_PATH.exists():
        raise FileNotFoundError(
            f"Audit tracker not found at: {TRACKER_PATH}"
        )

    return load_workbook(TRACKER_PATH)


def get_or_create_sheet(workbook, sheet_name: str):
    """Return the matching worksheet or create it if missing."""

    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]

    return workbook.create_sheet(title=sheet_name)


def ensure_headers(
    worksheet,
    headers: list[str],
) -> None:
    """Ensure the worksheet headers are always stored in row 1."""

    current_first_row = [
        worksheet.cell(row=1, column=index).value
        for index in range(1, len(headers) + 1)
    ]

    # Headers already exist correctly.
    if current_first_row == headers:
        return

    first_row_is_empty = all(
        value is None
        for value in current_first_row
    )

    # If row 1 already contains data, create a new header row.
    if not first_row_is_empty:
        worksheet.insert_rows(1)

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

        cell.font = Font(bold=True)

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )

    worksheet.freeze_panes = "A2"


def get_existing_issue_numbers(
    worksheet,
    issue_column_name: str = "Issue",
) -> set[str]:
    """Read existing issue IDs from the hyperlink formula column."""

    headers = [cell.value for cell in worksheet[1]]

    if issue_column_name not in headers:
        return set()

    issue_column_index = headers.index(issue_column_name) + 1
    existing_issue_ids = set()

    for row_number in range(2, worksheet.max_row + 1):
        value = worksheet.cell(
            row=row_number,
            column=issue_column_index,
        ).value

        if not value:
            continue

        match = re.search(r'"(ISSUE-\d+)"\)$', str(value))

        if match:
            existing_issue_ids.add(match.group(1))

    return existing_issue_ids


def insert_dataframe_at_top(
    worksheet,
    dataframe: pd.DataFrame,
) -> int:
    """
    Insert new records directly below the header row.

    Existing rows are shifted down so the newest issues appear first.
    """

    if dataframe.empty:
        return 0

    number_of_rows = len(dataframe)

    worksheet.insert_rows(
        idx=2,
        amount=number_of_rows,
    )

    for row_index, row_values in enumerate(
        dataframe.itertuples(index=False, name=None),
        start=2,
    ):
        for column_index, value in enumerate(
            row_values,
            start=1,
        ):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    return number_of_rows


def save_tracker(workbook) -> None:
    """Save the updated workbook."""

    workbook.save(TRACKER_PATH)